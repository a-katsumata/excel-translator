import openpyxl
import deepl
import os
import re

# --- 設定 (ここをご自身の環境に合わせて編集してください) ---

# 1. DeepLのAPIキーを入力してください
DEEPL_API_KEY = "a8ee58ad-8642-4c06-85b4-bc7d0e6e35a8:fx"

# 2. 翻訳したい元のExcelファイル名
INPUT_FILENAME = "日程表.xlsx"

# 3. 翻訳後に保存するファイル名
OUTPUT_FILENAME = "Itinerary_English.xlsx"

# 4. 日本語の定型文を、より翻訳しやすい表現に置換するルール
ITINERARY_REPLACEMENTS = {
    "食事：": "Meal: ",
    "朝食": "Breakfast",
    "昼食": "Lunch",
    "夕食": "Dinner",
    "宿泊：": "Accommodation: ",
    "様": "",  # 「田中様」を「Tanaka」とするため
    "ご一行": "Group",
    "各自": "on your own / at your leisure",
    "自由行動": "Free time",
    "---": "---" # 区切り線はそのまま
}
# -----------------------------------------------------------


def preprocess_text(text, replacements):
    """
    翻訳前に、定義したルールに基づいてテキストを置換する関数
    """
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text

def translate_excel_itinerary():
    """
    旅行の旅程表Excelを、文脈を考慮して自然な英語に翻訳する。
    """
    if "YOUR_DEEPL_API_KEY" in DEEPL_API_KEY:
        print("エラー: スクリプト内の DEEPL_API_KEY をご自身のキーに書き換えてください。")
        return
    if not os.path.exists(INPUT_FILENAME):
        print(f"エラー: 入力ファイル '{INPUT_FILENAME}' が見つかりません。")
        return

    try:
        translator = deepl.Translator(DEEPL_API_KEY)
        print("DeepL APIに接続しました。")

        workbook = openpyxl.load_workbook(INPUT_FILENAME)
        print(f"'{INPUT_FILENAME}' を読み込み、旅程表として翻訳を開始します...")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\nシート '{sheet_name}' の翻訳を開始...")

            cells_to_translate = []
            texts_to_translate = []

            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip():
                        cells_to_translate.append(cell)
                        # --- 前処理の実行 ---
                        text = preprocess_text(cell.value, ITINERARY_REPLACEMENTS)
                        texts_to_translate.append(text)

            if not texts_to_translate:
                print(f"シート '{sheet_name}' に翻訳対象のテキストがありませんでした。")
                continue

            # --- DeepL APIで翻訳を実行 ---
            results = translator.translate_text(
                texts_to_translate,
                source_lang="JA",
                target_lang="EN-US", # または "EN-GB"
                # ★改善点: 文脈を伝える
                context="This is a travel itinerary for a tour group."
            )

            # 翻訳結果をセルに書き戻す
            for cell, result in zip(cells_to_translate, results):
                print(f"  {cell.coordinate}: '{cell.value}' -> '{result.text}'")
                cell.value = result.text

            print(f"シート '{sheet_name}' の翻訳が完了しました。")

        workbook.save(OUTPUT_FILENAME)
        print(f"\n✅ 翻訳がすべて完了しました！ 結果を '{OUTPUT_FILENAME}' に保存しました。")

    except deepl.exceptions.AuthorizationError:
        print("エラー: DeepLのAPIキーが無効です。")
    except deepl.exceptions.QuotaExceededException:
        print("エラー: DeepLの無料版の文字数制限に達しました。")
    except Exception as e:
        print(f"予期せぬエラーが発生しました: {e}")


if __name__ == "__main__":
    translate_excel_itinerary()
