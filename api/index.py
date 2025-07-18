from flask import Flask, jsonify, request, render_template, send_file
import os
import sys
import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy as xlutils_copy
import requests
import io
import tempfile
from urllib.parse import quote
import re
import gc
import sys
from datetime import datetime

# パスを追加
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.insert(0, parent_dir)

app = Flask(__name__, template_folder='../templates')
app.secret_key = os.environ.get('SECRET_KEY', 'excel-translator-secret-key')

def should_translate_cell(cell_value):
    """セルの内容を分析して翻訳が必要かどうかを判定"""
    if not cell_value:
        return False
    
    # 文字列以外は翻訳しない
    if not isinstance(cell_value, str):
        return False
    
    # 空白文字のみは翻訳しない
    if not cell_value.strip():
        return False
    
    # 数値のみの場合は翻訳しない
    if re.match(r'^[\d\s,.\-+%$€¥]+$', cell_value.strip()):
        return False
    
    # 日付形式の場合は翻訳しない
    date_patterns = [
        r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}$',  # 2023-12-31, 2023/12/31
        r'^\d{1,2}[-/]\d{1,2}[-/]\d{4}$',  # 31-12-2023, 31/12/2023
        r'^\d{4}年\d{1,2}月\d{1,2}日$',     # 2023年12月31日
    ]
    
    for pattern in date_patterns:
        if re.match(pattern, cell_value.strip()):
            return False
    
    # 数式の場合は翻訳しない
    if cell_value.startswith('='):
        return False
    
    # URLの場合は翻訳しない
    if re.match(r'^https?://', cell_value.strip()):
        return False
    
    # メールアドレスの場合は翻訳しない
    if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', cell_value.strip()):
        return False
    
    # 長すぎる場合は翻訳しない（API制限回避）
    if len(cell_value) > 5000:
        return False
    
    # 短すぎる場合（1文字）で日本語/中国語/韓国語でない場合は翻訳しない
    if len(cell_value.strip()) == 1:
        if not re.match(r'[ひらがなカタカナ漢字가-힣]', cell_value):
            return False
    
    return True

def generate_context_from_headers(sheet, cell_row, cell_col):
    """ヘッダー情報から文脈を生成"""
    context_parts = []
    
    # 同じ行の左側のセルから文脈を取得（見出し）
    for col in range(max(1, cell_col - 3), cell_col):
        if col < cell_col:
            header_cell = sheet.cell(row=cell_row, column=col)
            if header_cell.value and isinstance(header_cell.value, str):
                context_parts.append(header_cell.value)
    
    # 同じ列の上側のセルから文脈を取得（カラムヘッダー）
    for row in range(max(1, cell_row - 3), cell_row):
        if row < cell_row:
            header_cell = sheet.cell(row=row, column=cell_col)
            if header_cell.value and isinstance(header_cell.value, str):
                context_parts.append(header_cell.value)
    
    return ' '.join(context_parts[:5])  # 最大5つの要素で文脈を作成

def analyze_sheet_structure(sheet):
    """シート構造を分析して翻訳単位を決定"""
    # シートの全セルを取得
    all_cells = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                all_cells.append({
                    'row': cell.row,
                    'column': cell.column,
                    'value': cell.value,
                    'coordinate': cell.coordinate
                })
    
    # ヘッダー行を特定（最初の数行でテキストが多い行）
    header_rows = []
    for row_num in range(1, min(6, sheet.max_row + 1)):
        text_cells = 0
        for col_num in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 0:
                text_cells += 1
        if text_cells >= sheet.max_column * 0.5:  # 50%以上がテキスト
            header_rows.append(row_num)
    
    # データ領域を特定
    data_start_row = max(header_rows) + 1 if header_rows else 1
    
    return {
        'header_rows': header_rows,
        'data_start_row': data_start_row,
        'total_cells': len(all_cells),
        'max_row': sheet.max_row,
        'max_column': sheet.max_column
    }

def create_sheet_context(sheet, structure):
    """シート全体の文脈を作成"""
    context_parts = []
    
    # シート名を文脈に追加
    if sheet.title:
        context_parts.append(f"シート名: {sheet.title}")
    
    # ヘッダー行から文脈を作成
    for header_row in structure['header_rows']:
        header_texts = []
        for col_num in range(1, min(sheet.max_column + 1, 10)):  # 最大10列
            cell = sheet.cell(row=header_row, column=col_num)
            if cell.value and isinstance(cell.value, str):
                header_texts.append(cell.value)
        if header_texts:
            context_parts.append(' | '.join(header_texts))
    
    return '. '.join(context_parts)

def create_cell_mapping(sheet):
    """セルの位置と内容のマッピングを作成"""
    cell_mapping = {}
    translation_tasks = []
    
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell_key = f"{row}_{col}"
            
            # セルの情報を保存
            cell_mapping[cell_key] = {
                'row': row,
                'col': col,
                'coordinate': cell.coordinate,
                'original_value': cell.value,
                'needs_translation': should_translate_cell(cell.value),
                'cell_object': cell
            }
            
            # 翻訳が必要なセルを翻訳タスクに追加
            if should_translate_cell(cell.value):
                translation_tasks.append({
                    'cell_key': cell_key,
                    'text': str(cell.value),
                    'context': generate_context_from_headers(sheet, row, col)
                })
    
    return cell_mapping, translation_tasks

def calculate_text_size(texts):
    """テキストリストの合計文字数を計算"""
    return sum(len(str(text)) for text in texts)

def estimate_payload_size(texts):
    """JSONペイロードの推定サイズを計算（オーバーヘッドを含む）"""
    # 基本的なJSON構造のオーバーヘッド
    base_overhead = 200  # リクエストヘッダーなど
    
    # 各テキストのJSONオーバーヘッド（引用符、カンマ、構造など）
    text_overhead = sum(len(str(text)) * 1.2 + 20 for text in texts)  # 20%のオーバーヘッド + 固定値
    
    return int(base_overhead + text_overhead)

def create_dynamic_batches(translation_tasks, max_chars_per_batch=50000):
    """文字数制限とJSONペイロードサイズに基づいて動的にバッチを作成"""
    batches = []
    current_batch = []
    current_char_count = 0
    
    # 安全マージンを追加（実際の制限の80%を使用）
    safe_limit = int(max_chars_per_batch * 0.8)
    
    for task in translation_tasks:
        text_length = len(str(task['text']))
        
        # 単一のセルが制限を超える場合は個別処理
        if text_length > safe_limit:
            # 現在のバッチがあれば追加
            if current_batch:
                batches.append(current_batch)
                current_batch = []
                current_char_count = 0
            
            # 長いセルは個別バッチとして処理
            batches.append([task])
            continue
        
        # バッチに追加すると制限を超える場合をチェック
        test_batch = current_batch + [task]
        test_texts = [t['text'] for t in test_batch]
        estimated_payload = estimate_payload_size(test_texts)
        
        if estimated_payload > safe_limit and current_batch:
            # 現在のバッチを完成させる
            batches.append(current_batch)
            current_batch = [task]
            current_char_count = text_length
        else:
            # バッチに追加
            current_batch.append(task)
            current_char_count += text_length
    
    # 最後のバッチを追加
    if current_batch:
        batches.append(current_batch)
    
    return batches

def translate_with_staged_fallback(translation_tasks, sheet, context, target_lang, source_lang, formality, api_key, processing_params):
    """段階的フォールバック処理付きの翻訳"""
    if not translation_tasks:
        return {}
    
    # 処理パラメータを取得
    max_chars_per_batch = processing_params['max_chars_per_batch']
    context_limit = processing_params['context_limit']
    enable_fallback = processing_params['enable_fallback']
    
    # 文脈の最適化
    sheet_context = f"シート名: {sheet.title}. " if sheet.title else ""
    
    # ヘッダー情報の簡潔化
    header_info = []
    for row in range(1, min(3, sheet.max_row + 1)):
        row_texts = []
        for col in range(1, min(sheet.max_column + 1, 8)):
            cell = sheet.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and len(str(cell.value)) < 50:
                row_texts.append(str(cell.value))
        if row_texts:
            header_info.append(" | ".join(row_texts))
    
    if header_info:
        sheet_context += "ヘッダー情報: " + "; ".join(header_info[:2]) + ". "
    
    # 文脈の長さ制限
    full_context = f"{context}. {sheet_context}" if context else sheet_context
    if len(full_context) > context_limit:
        full_context = full_context[:context_limit] + "..."
    
    # 動的バッチ作成
    batches = create_dynamic_batches(translation_tasks, max_chars_per_batch)
    translations = {}
    failed_tasks = []
    
    print(f"Processing {len(translation_tasks)} tasks in {len(batches)} batches")
    
    # 第1段階: 通常のバッチ処理
    for batch_idx, batch_tasks in enumerate(batches):
        batch_texts = [task['text'] for task in batch_tasks]
        batch_char_count = calculate_text_size(batch_texts)
        
        print(f"Batch {batch_idx + 1}/{len(batches)}: {len(batch_tasks)} tasks, {batch_char_count} chars")
        
        try:
            translated_batch = translate_batch(
                batch_texts,
                target_lang,
                source_lang,
                full_context,
                api_key,
                formality
            )
            
            # 翻訳結果をマッピング
            for j, task in enumerate(batch_tasks):
                if j < len(translated_batch):
                    translations[task['cell_key']] = translated_batch[j]
                else:
                    failed_tasks.append(task)
                    
        except Exception as e:
            error_msg = str(e)
            print(f"Translation batch {batch_idx + 1} error: {error_msg}")
            
            # 413エラー（Payload too large）の場合、バッチを分割して再試行
            if "413" in error_msg and "Payload too large" in error_msg:
                print(f"Payload too large error detected. Splitting batch of {len(batch_tasks)} tasks...")
                
                # バッチを半分に分割して再試行
                mid_point = len(batch_tasks) // 2
                if mid_point > 0:
                    first_half = batch_tasks[:mid_point]
                    second_half = batch_tasks[mid_point:]
                    
                    # 第1半分を処理
                    try:
                        first_texts = [task['text'] for task in first_half]
                        first_translated = translate_batch(
                            first_texts, target_lang, source_lang, formality, 
                            deepl_api_key, full_context
                        )
                        for j, task in enumerate(first_half):
                            if j < len(first_translated):
                                translations[task['cell_key']] = first_translated[j]
                            else:
                                failed_tasks.append(task)
                        print(f"First half ({len(first_half)} tasks) processed successfully")
                    except Exception as e2:
                        print(f"First half still failed: {str(e2)}")
                        failed_tasks.extend(first_half)
                    
                    # 第2半分を処理
                    try:
                        second_texts = [task['text'] for task in second_half]
                        second_translated = translate_batch(
                            second_texts, target_lang, source_lang, formality, 
                            deepl_api_key, full_context
                        )
                        for j, task in enumerate(second_half):
                            if j < len(second_translated):
                                translations[task['cell_key']] = second_translated[j]
                            else:
                                failed_tasks.append(task)
                        print(f"Second half ({len(second_half)} tasks) processed successfully")
                    except Exception as e3:
                        print(f"Second half still failed: {str(e3)}")
                        failed_tasks.extend(second_half)
                else:
                    # 単一タスクでも413エラーの場合は失敗として扱う
                    failed_tasks.extend(batch_tasks)
            else:
                # 413エラー以外の場合は通常の失敗として扱う
                failed_tasks.extend(batch_tasks)
        
        # メモリ解放
        del batch_texts
        gc.collect()
    
    # 第2段階: 失敗したタスクの個別処理（フォールバック有効時）
    if failed_tasks and enable_fallback:
        print(f"Fallback processing for {len(failed_tasks)} failed tasks")
        
        for task in failed_tasks:
            try:
                # 文脈を簡略化して個別処理
                simple_context = context[:200] if context else ""
                single_translation = translate_batch(
                    [task['text']],
                    target_lang,
                    source_lang,
                    simple_context,
                    api_key,
                    formality
                )
                
                if single_translation:
                    translations[task['cell_key']] = single_translation[0]
                else:
                    translations[task['cell_key']] = task['text']
                    
            except Exception as single_error:
                print(f"Single task fallback error: {str(single_error)}")
                translations[task['cell_key']] = task['text']
    
    # 第3段階: 文脈なしでの最終試行
    remaining_failed = []
    for task in failed_tasks:
        if task['cell_key'] not in translations:
            remaining_failed.append(task)
    
    if remaining_failed and enable_fallback:
        print(f"Final fallback processing for {len(remaining_failed)} tasks")
        
        for task in remaining_failed:
            try:
                # 文脈なしで処理
                final_translation = translate_batch(
                    [task['text']],
                    target_lang,
                    source_lang,
                    "",
                    api_key,
                    formality
                )
                
                if final_translation:
                    translations[task['cell_key']] = final_translation[0]
                else:
                    translations[task['cell_key']] = task['text']
                    
            except Exception as final_error:
                print(f"Final fallback error: {str(final_error)}")
                translations[task['cell_key']] = task['text']
    
    return translations


def apply_translations_to_sheet(sheet, cell_mapping, translations):
    """翻訳結果をシートに適用"""
    for cell_key, translation in translations.items():
        if cell_key in cell_mapping:
            cell_info = cell_mapping[cell_key]
            if cell_info['needs_translation']:
                cell_info['cell_object'].value = translation

def preserve_merged_cells(sheet):
    """結合セルの情報を保存"""
    merged_ranges = []
    try:
        if hasattr(sheet, 'merged_cells') and hasattr(sheet.merged_cells, 'ranges'):
            # XLSXの場合
            for merged_range in sheet.merged_cells.ranges:
                merged_ranges.append(str(merged_range))
        elif hasattr(sheet, 'merged_cells') and isinstance(sheet.merged_cells, list):
            # XLSの場合
            for merged_range in sheet.merged_cells:
                merged_ranges.append(merged_range)
    except:
        pass
    return merged_ranges

def restore_merged_cells(sheet, merged_ranges):
    """結合セルの情報を復元"""
    if not merged_ranges:
        return
        
    for merged_range in merged_ranges:
        try:
            # XLSXの場合は文字列から復元
            if isinstance(merged_range, str):
                sheet.merge_cells(merged_range)
            # XLSの場合はタプルから復元
            elif isinstance(merged_range, (list, tuple)) and len(merged_range) == 4:
                r1, r2, c1, c2 = merged_range
                sheet.merge_cells(start_row=r1+1, start_column=c1+1, 
                                end_row=r2, end_column=c2)
        except Exception as e:
            print(f"Failed to restore merged cell {merged_range}: {str(e)}")

def validate_translation_accuracy(sheet, cell_mapping, translations):
    """翻訳の正確性を検証"""
    validation_results = {
        'total_cells': len(cell_mapping),
        'cells_needing_translation': 0,
        'cells_translated': 0,
        'cells_preserved': 0,
        'errors': []
    }
    
    for cell_key, cell_info in cell_mapping.items():
        if cell_info['needs_translation']:
            validation_results['cells_needing_translation'] += 1
            
            if cell_key in translations:
                validation_results['cells_translated'] += 1
                
                # 翻訳結果の妥当性チェック
                original = cell_info['original_value']
                translated = translations[cell_key]
                
                # 明らかに不適切な翻訳の検出
                if len(str(translated)) == 0 and len(str(original)) > 0:
                    validation_results['errors'].append(f"Empty translation for cell {cell_info['coordinate']}")
                elif len(str(translated)) > len(str(original)) * 10:
                    validation_results['errors'].append(f"Suspiciously long translation for cell {cell_info['coordinate']}")
            else:
                validation_results['errors'].append(f"Missing translation for cell {cell_info['coordinate']}")
        else:
            validation_results['cells_preserved'] += 1
    
    return validation_results

def analyze_file_complexity(wb):
    """ファイルの複雑さを分析して処理戦略を決定"""
    analysis = {
        'total_sheets': len(wb.sheetnames),
        'total_cells': 0,
        'total_text_chars': 0,
        'max_sheet_cells': 0,
        'has_merged_cells': False,
        'complexity_score': 0,
        'processing_strategy': 'standard'
    }
    
    for sheet_name in wb.sheetnames:
        sheet = wb.get_sheet(sheet_name)
        sheet_cells = 0
        sheet_text_chars = 0
        
        # 結合セルの確認
        if len(sheet.merged_cells.ranges) > 0:
            analysis['has_merged_cells'] = True
        
        # セルの分析
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    sheet_cells += 1
                    if isinstance(cell.value, str):
                        sheet_text_chars += len(cell.value)
        
        analysis['total_cells'] += sheet_cells
        analysis['total_text_chars'] += sheet_text_chars
        analysis['max_sheet_cells'] = max(analysis['max_sheet_cells'], sheet_cells)
    
    # 複雑さスコアの計算
    analysis['complexity_score'] = (
        analysis['total_cells'] * 0.1 +
        analysis['total_text_chars'] * 0.001 +
        analysis['total_sheets'] * 10 +
        (50 if analysis['has_merged_cells'] else 0)
    )
    
    # 処理戦略の決定（より保守的な基準）
    if analysis['complexity_score'] < 200:
        analysis['processing_strategy'] = 'fast'
    elif analysis['complexity_score'] < 800:
        analysis['processing_strategy'] = 'standard'
    else:
        analysis['processing_strategy'] = 'careful'
    
    # 大きなファイルの場合は強制的に慎重な処理
    if analysis['total_text_chars'] > 100000 or analysis['total_cells'] > 3000:
        analysis['processing_strategy'] = 'careful'
        print(f"Large file detected: {analysis['total_text_chars']} chars, {analysis['total_cells']} cells - using careful strategy")
    
    # 超大型ファイルの場合は超安全モード
    if analysis['total_text_chars'] > 300000 or analysis['total_cells'] > 8000:
        analysis['processing_strategy'] = 'ultra_safe'
        print(f"Very large file detected: {analysis['total_text_chars']} chars, {analysis['total_cells']} cells - using ultra_safe strategy")
    
    return analysis

def get_processing_parameters(strategy):
    """処理戦略に基づいてパラメータを設定"""
    params = {
        'fast': {
            'max_chars_per_batch': 15000,  # 大幅削減: 80000 → 15000
            'max_batches_per_sheet': 100,
            'context_limit': 1500,
            'enable_fallback': True  # フォールバック機能を有効化
        },
        'standard': {
            'max_chars_per_batch': 10000,  # 大幅削減: 50000 → 10000
            'max_batches_per_sheet': 200,
            'context_limit': 1000,
            'enable_fallback': True
        },
        'careful': {
            'max_chars_per_batch': 5000,   # 大幅削減: 30000 → 5000
            'max_batches_per_sheet': 500,
            'context_limit': 500,
            'enable_fallback': True
        },
        'ultra_safe': {
            'max_chars_per_batch': 2000,   # 超安全モード
            'max_batches_per_sheet': 1000,
            'context_limit': 200,
            'enable_fallback': True
        }
    }
    
    return params.get(strategy, params['standard'])

def detect_file_format(file_data):
    """ファイル形式（XLS/XLSX）を検出"""
    # ファイルの先頭バイトを確認
    file_data.seek(0)
    header = file_data.read(8)
    file_data.seek(0)
    
    # XLSX形式（ZIP形式）の場合
    if header.startswith(b'PK'):
        return 'xlsx'
    
    # XLS形式（OLE2形式）の場合
    if header.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'):
        return 'xls'
    
    # ファイル名から判定
    return 'unknown'

class UnifiedWorkbook:
    """XLS/XLSX両対応の統一ワークブッククラス"""
    
    def __init__(self, file_data, file_format):
        self.file_format = file_format
        self.original_filename = None
        self.translated_data = {}  # 翻訳データを保存
        self.original_file_data = file_data.getvalue()  # 元のファイルデータを保存
        
        if file_format == 'xlsx':
            self.workbook = openpyxl.load_workbook(file_data)
            self.sheetnames = self.workbook.sheetnames
        elif file_format == 'xls':
            self.workbook = xlrd.open_workbook(file_contents=file_data.read(), formatting_info=True)
            self.sheetnames = self.workbook.sheet_names()
            # XLSファイルの書き込み用ワークブックを作成
            self.write_workbook = None
        else:
            raise ValueError(f"Unsupported file format: {file_format}")
    
    def get_sheet(self, sheet_name):
        """シートを取得"""
        if self.file_format == 'xlsx':
            return UnifiedWorksheet(self.workbook[sheet_name], 'xlsx', self)
        elif self.file_format == 'xls':
            sheet_index = self.sheetnames.index(sheet_name)
            return UnifiedWorksheet(self.workbook.sheet_by_index(sheet_index), 'xls', self)
    
    def save(self, file_path):
        """ファイルを保存（元の形式を保持）"""
        if self.file_format == 'xlsx':
            self.workbook.save(file_path)
        elif self.file_format == 'xls':
            # XLSファイルを元の形式で保存（完全な形式保持）
            self._save_xls_with_translation(file_path)
    
    def _save_xls_with_translation(self, file_path):
        """XLSファイルを翻訳データと共に保存"""
        try:
            print(f"DEBUG: Starting XLS save with {len(self.translated_data)} translations")
            
            # xlutilsを使用して元のワークブックをコピー（すべての書式・図形を保持）
            self.write_workbook = xlutils_copy(self.workbook)
            print("DEBUG: xlutils copy completed")
            
            # 列幅と行高さの設定を保持
            self._preserve_column_row_dimensions()
            print("DEBUG: Column/row dimensions preserved")
            
            # 翻訳されたデータを書き込み
            for sheet_name in self.sheetnames:
                sheet_index = self.sheetnames.index(sheet_name)
                write_sheet = self.write_workbook.get_sheet(sheet_index)
                original_sheet = self.workbook.sheet_by_index(sheet_index)
                print(f"DEBUG: Processing sheet {sheet_name} (index {sheet_index})")
                
                # 翻訳データを適用
                translations_applied = 0
                for cell_key, translated_value in self.translated_data.items():
                    if cell_key.startswith(f"{sheet_name}_"):
                        parts = cell_key.split('_')
                        if len(parts) >= 3:
                            row = int(parts[-2]) - 1  # 0ベースに変換
                            col = int(parts[-1]) - 1  # 0ベースに変換
                            
                            try:
                                # 翻訳された値を書き込み（シンプルな方法）
                                write_sheet.write(row, col, translated_value)
                                translations_applied += 1
                                
                            except Exception as e:
                                print(f"Error writing translated value to cell {cell_key}: {e}")
                                # エラーが発生した場合はスキップ
                                continue
                
                print(f"DEBUG: Applied {translations_applied} translations to sheet {sheet_name}")
            
            # ファイルを保存
            print(f"DEBUG: Saving file to {file_path}")
            self.write_workbook.save(file_path)
            print("DEBUG: XLS file saved successfully")
            
        except Exception as e:
            print(f"ERROR in _save_xls_with_translation: {e}")
            import traceback
            traceback.print_exc()
            raise
    
    def _preserve_column_row_dimensions(self):
        """列幅と行高さの設定を保持"""
        if not hasattr(self, 'write_workbook') or self.write_workbook is None:
            return
        
        try:
            print("DEBUG: Starting column/row dimension preservation")
            # 元のワークブックから列幅と行高さの情報を取得して適用
            for sheet_index in range(len(self.sheetnames)):
                original_sheet = self.workbook.sheet_by_index(sheet_index)
                write_sheet = self.write_workbook.get_sheet(sheet_index)
                
                # 列幅の設定を保持（基本的なバージョン）
                try:
                    if hasattr(original_sheet, 'colinfo_map') and original_sheet.colinfo_map:
                        for col_index, col_info in original_sheet.colinfo_map.items():
                            if col_info and hasattr(col_info, 'width'):
                                try:
                                    # xlwtの列幅設定（簡単な方法）
                                    width = int(col_info.width)
                                    write_sheet.col(col_index).width = width
                                except:
                                    pass
                except Exception as e:
                    print(f"DEBUG: Column width preservation failed: {e}")
                
                # 行高さの設定を保持（基本的なバージョン）
                try:
                    if hasattr(original_sheet, 'rowinfo_map') and original_sheet.rowinfo_map:
                        for row_index, row_info in original_sheet.rowinfo_map.items():
                            if row_info and hasattr(row_info, 'height'):
                                try:
                                    # xlwtの行高さ設定（簡単な方法）
                                    height = int(row_info.height)
                                    write_sheet.row(row_index).height = height
                                except:
                                    pass
                except Exception as e:
                    print(f"DEBUG: Row height preservation failed: {e}")
            
            print("DEBUG: Column/row dimension preservation completed")
        except Exception as e:
            print(f"Warning: Could not preserve column/row dimensions: {e}")
            # エラーが発生しても処理を続行

class UnifiedWorksheet:
    """XLS/XLSX両対応の統一ワークシートクラス"""
    
    def __init__(self, sheet, file_format, workbook=None):
        self.sheet = sheet
        self.file_format = file_format
        self.workbook = workbook
        self.title = sheet.title if file_format == 'xlsx' else sheet.name
        
        if file_format == 'xlsx':
            self.max_row = sheet.max_row
            self.max_column = sheet.max_column
            self.merged_cells = sheet.merged_cells
        elif file_format == 'xls':
            self.max_row = sheet.nrows
            self.max_column = sheet.ncols
            # XLSの結合セル情報を取得
            self.merged_cells = self._get_xls_merged_cells()
    
    def _get_xls_merged_cells(self):
        """XLSの結合セル情報を取得"""
        merged_cells = DummyMergedCells()
        try:
            if hasattr(self.sheet, 'merged_cells'):
                merged_cells.ranges = self.sheet.merged_cells
        except:
            pass
        return merged_cells
    
    def cell(self, row, column):
        """セルを取得"""
        if self.file_format == 'xlsx':
            return self.sheet.cell(row=row, column=column)
        elif self.file_format == 'xls':
            return UnifiedCell(self.sheet, row-1, column-1, 'xls', self.workbook, self.title)
    
    def iter_rows(self):
        """行をイテレート"""
        for row in range(1, self.max_row + 1):
            yield [self.cell(row, col) for col in range(1, self.max_column + 1)]

class UnifiedCell:
    """XLS/XLSX両対応の統一セルクラス"""
    
    def __init__(self, sheet, row, column, file_format='xls', workbook=None, sheet_name=None):
        self.sheet = sheet
        self.row = row + 1  # 1ベースに変換
        self.column = column + 1  # 1ベースに変換
        self.coordinate = f"{chr(65 + column)}{row + 1}"
        self.file_format = file_format
        self.workbook = workbook
        self.sheet_name = sheet_name
        self._value = None
        
        if file_format == 'xls':
            try:
                self._value = sheet.cell_value(row, column)
                if self._value == '':
                    self._value = None
            except:
                self._value = None
    
    @property
    def value(self):
        # XLSファイルの場合、翻訳されたデータがあればそれを返す
        if self.file_format == 'xls' and self.workbook:
            cell_key = f"{self.sheet_name}_{self.row}_{self.column}"
            if cell_key in self.workbook.translated_data:
                return self.workbook.translated_data[cell_key]
        return self._value
    
    @value.setter
    def value(self, new_value):
        if self.file_format == 'xls' and self.workbook:
            # XLSファイルの場合、翻訳データをworkbookに保存
            cell_key = f"{self.sheet_name}_{self.row}_{self.column}"
            self.workbook.translated_data[cell_key] = new_value
        else:
            self._value = new_value

class DummyMergedCells:
    """XLS用のダミー結合セルクラス"""
    
    def __init__(self):
        self.ranges = []
    
    def __len__(self):
        return len(self.ranges)
    
    def __iter__(self):
        return iter(self.ranges)

@app.route('/')
def index():
    try:
        return render_template('index.html')
    except Exception as e:
        return jsonify({
            'error': 'Template error',
            'message': str(e),
            'working_directory': os.getcwd(),
            'template_folder': app.template_folder,
            'files_in_templates': os.listdir(os.path.join(parent_dir, 'templates')) if os.path.exists(os.path.join(parent_dir, 'templates')) else 'templates directory not found'
        }), 500

@app.route('/health')
def health():
    return jsonify({
        'status': 'healthy',
        'service': 'excel-translator',
        'python_version': sys.version,
        'working_directory': os.getcwd(),
        'parent_directory': parent_dir,
        'current_directory': current_dir,
        'template_folder': app.template_folder,
        'environment_variables': list(os.environ.keys()),
        'deepl_api_key_exists': bool(os.environ.get('DEEPL_API_KEY')),
        'files_in_current_dir': os.listdir(os.getcwd()),
        'files_in_parent_dir': os.listdir(parent_dir) if os.path.exists(parent_dir) else 'parent directory not found'
    })

def translate_batch(texts, target_lang, source_lang, context, api_key, formality=None):
    """DeepL APIを使用して複数のテキストを一括翻訳"""
    if not texts:
        return []
    
    # 空のテキストを除外し、インデックスを記録
    non_empty_texts = []
    text_indices = []
    
    for i, text in enumerate(texts):
        if text and text.strip():
            non_empty_texts.append(text)
            text_indices.append(i)
    
    if not non_empty_texts:
        return texts
    
    url = "https://api-free.deepl.com/v2/translate"
    
    data = {
        'auth_key': api_key,
        'text': non_empty_texts,
        'target_lang': target_lang,
        'source_lang': source_lang if source_lang != 'auto' else None
    }
    
    if context:
        data['context'] = context
    
    # フォーマリティの設定
    if formality and formality != 'default':
        data['formality'] = formality
    
    # 常に高品質モードを使用
    data['model_type'] = 'quality_optimized'
    
    response = requests.post(url, data=data)
    
    if response.status_code == 200:
        result = response.json()
        translated_texts = [t['text'] for t in result['translations']]
        
        # 結果を元の配列に戻す
        final_results = list(texts)
        for i, translated_text in enumerate(translated_texts):
            final_results[text_indices[i]] = translated_text
        
        return final_results
    else:
        raise Exception(f"DeepL API error: {response.status_code} - {response.text}")

@app.route('/api/translate', methods=['POST'])
def api_translate():
    try:
        # 環境変数チェック
        deepl_api_key = os.environ.get('DEEPL_API_KEY')
        if not deepl_api_key:
            return jsonify({'error': 'DEEPL_API_KEY not found in environment variables'}), 500
        
        # ファイルチェック
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # パラメータ取得
        source_lang = request.form.get('source_lang', 'JA')
        target_lang = request.form.get('target_lang', 'EN-US')
        context = request.form.get('context', '')
        formality = request.form.get('formality', 'default')
        
        # ファイル形式を検出
        file_data = io.BytesIO(file.read())
        file_format = detect_file_format(file_data)
        
        if file_format == 'unknown':
            # ファイル名から判定
            file_extension = file.filename.lower().split('.')[-1] if '.' in file.filename else ''
            if file_extension == 'xlsx':
                file_format = 'xlsx'
            elif file_extension == 'xls':
                file_format = 'xls'
            else:
                return jsonify({'error': 'Unsupported file format. Please use .xlsx or .xls files.'}), 400
        
        print(f"Detected file format: {file_format}")
        
        # 統一ワークブックを作成
        try:
            wb = UnifiedWorkbook(file_data, file_format)
            print(f"Successfully created UnifiedWorkbook with {len(wb.sheetnames)} sheets")
        except Exception as e:
            print(f"Error creating UnifiedWorkbook: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Failed to read file: {str(e)}'}), 500
        
        # ファイルの複雑さを分析
        file_analysis = analyze_file_complexity(wb)
        processing_params = get_processing_parameters(file_analysis['processing_strategy'])
        
        print(f"File analysis: {file_analysis['total_sheets']} sheets, {file_analysis['total_cells']} cells, {file_analysis['total_text_chars']} chars")
        print(f"Processing strategy: {file_analysis['processing_strategy']}")
        print(f"Processing parameters: {processing_params}")
        
        # 全シートを新しいセル対応保証アルゴリズムで処理
        for sheet_name in wb.sheetnames:
            sheet = wb.get_sheet(sheet_name)
            
            print(f"Processing sheet: {sheet_name}")
            
            # 結合セルの情報を保存
            merged_ranges = preserve_merged_cells(sheet)
            
            # セルマッピングと翻訳タスクを作成
            cell_mapping, translation_tasks = create_cell_mapping(sheet)
            
            if not translation_tasks:
                print(f"No translation tasks found for sheet {sheet_name}")
                continue
            
            # 翻訳の実行（段階的フォールバック付き）
            translations = translate_with_staged_fallback(
                translation_tasks,
                sheet,
                context,
                target_lang,
                source_lang,
                formality,
                deepl_api_key,
                processing_params
            )
            
            # 翻訳結果をシートに適用
            apply_translations_to_sheet(sheet, cell_mapping, translations)
            
            # 翻訳の正確性を検証
            validation_results = validate_translation_accuracy(sheet, cell_mapping, translations)
            if validation_results['errors']:
                print(f"Validation errors for sheet {sheet_name}: {validation_results['errors']}")
            
            print(f"Sheet {sheet_name} completed: {validation_results['cells_translated']}/{validation_results['cells_needing_translation']} cells translated")
            
            # 結合セルを復元
            restore_merged_cells(sheet, merged_ranges)
            
            # シート処理後のメモリ解放
            gc.collect()
        
        # 翻訳されたファイルを一時ファイルに保存（元の形式を保持）
        file_extension = '.xlsx' if wb.file_format == 'xlsx' else '.xls'
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_file:
                print(f"Saving translated file as {file_extension} format")
                wb.save(tmp_file.name)
                tmp_file_path = tmp_file.name
                print(f"Successfully saved to {tmp_file_path}")
        except Exception as e:
            print(f"Error saving translated file: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Failed to save translated file: {str(e)}'}), 500
        
        # ファイル名を生成（翻訳済みの接頭辞を追加、元の拡張子を保持）
        original_filename = file.filename
        name, ext = os.path.splitext(original_filename)
        translated_filename = f"{name}_translated{ext}"
        
        # ファイルをダウンロード用に送信（適切なMIMEタイプを設定）
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' if wb.file_format == 'xlsx' else 'application/vnd.ms-excel'
        
        return send_file(
            tmp_file_path,
            as_attachment=True,
            download_name=translated_filename,
            mimetype=mimetype
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Vercel用のエクスポート
def app_handler(environ, start_response):
    return app(environ, start_response)

# Vercel用のapp変数をエクスポート
if __name__ == '__main__':
    app.run(debug=False)