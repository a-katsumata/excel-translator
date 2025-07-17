import openpyxl
import deepl
import os
import io
import logging
from typing import Dict, Any, List, Optional
from functools import lru_cache

# ログ設定
logger = logging.getLogger(__name__)

class ExcelTranslator:
    """
    Excelファイルの翻訳を行うクラス
    セルの結合、フォーマット、構造を保持しながら翻訳を実行
    """
    
    def __init__(self, deepl_api_key: str):
        """
        翻訳クラスの初期化
        
        Args:
            deepl_api_key: DeepL APIキー
        """
        self.deepl_api_key = deepl_api_key
        self.translator = deepl.Translator(deepl_api_key)
        logger.info("ExcelTranslator initialized")
        
    @lru_cache(maxsize=32)
    def get_context_replacements(self, context: str) -> Dict[str, str]:
        """
        文脈に応じた前処理置換ルールを取得
        
        Args:
            context: 文脈（日程表、事業計画など）
            
        Returns:
            置換ルール辞書
        """
        context_lower = context.lower()
        
        if "日程" in context_lower or "itinerary" in context_lower or "schedule" in context_lower:
            return {
                "食事：": "Meal: ",
                "朝食": "Breakfast",
                "昼食": "Lunch",
                "夕食": "Dinner",
                "宿泊：": "Accommodation: ",
                "様": "",
                "ご一行": "Group",
                "各自": "on your own / at your leisure",
                "自由行動": "Free time",
                "---": "---"
            }
        elif "事業計画" in context_lower or "business plan" in context_lower:
            return {
                "売上": "Revenue",
                "利益": "Profit",
                "予算": "Budget",
                "計画": "Plan",
                "目標": "Target",
                "実績": "Actual",
                "前年比": "Year-on-year",
                "四半期": "Quarter"
            }
        elif "財務" in context_lower or "financial" in context_lower:
            return {
                "資産": "Assets",
                "負債": "Liabilities",
                "資本": "Capital",
                "収入": "Income",
                "支出": "Expenses",
                "残高": "Balance",
                "合計": "Total"
            }
        else:
            return {}
    
    def preprocess_text(self, text: str, replacements: Dict[str, str]) -> str:
        """
        翻訳前のテキスト前処理
        
        Args:
            text: 処理対象テキスト
            replacements: 置換ルール
            
        Returns:
            処理後テキスト
        """
        if not text or not isinstance(text, str):
            return text
            
        for old, new in replacements.items():
            text = text.replace(old, new)
        return text
    
    def should_translate_text(self, text: str) -> bool:
        """
        テキストが翻訳対象かどうかを判定
        
        Args:
            text: 判定対象のテキスト
            
        Returns:
            翻訳対象の場合True、そうでなければFalse
        """
        if not text or not isinstance(text, str):
            return False
            
        text = text.strip()
        
        # 空文字の場合は翻訳不要
        if not text:
            return False
            
        # 記号のみの場合は翻訳不要
        symbol_only_patterns = [
            r'^[○×△▲▼◆◇□■☆★※・]+$',  # 日本語記号
            r'^[✓✗✘✔×○◯△▲▼◆◇□■☆★※・]+$',  # 一般的な記号
            r'^[0-9\s\-\.\,\(\)\[\]]+$',  # 数字と基本記号のみ
            r'^[\s\-\.\,\(\)\[\]\/\\]+$',  # 区切り記号のみ
            r'^[A-Z0-9\s\-\.\,\(\)\[\]]+$',  # 英数字と基本記号のみ（短い場合）
        ]
        
        import re
        for pattern in symbol_only_patterns:
            if re.match(pattern, text):
                return False
                
        # 短い英数字のみの文字列は翻訳しない
        if len(text) <= 3 and re.match(r'^[A-Za-z0-9\s\-\.\,\(\)\[\]]+$', text):
            return False
            
        return True
    
    def translate_excel_file(self, file_data: bytes, context: str = "", 
                           source_lang: str = "JA", target_lang: str = "EN-US") -> bytes:
        """
        Excelファイルの翻訳を実行
        
        Args:
            file_data: Excelファイルのバイトデータ
            context: 翻訳文脈
            source_lang: 翻訳元言語
            target_lang: 翻訳先言語
            
        Returns:
            翻訳後のExcelファイルバイトデータ
        """
        try:
            logger.info(f"Starting translation: {source_lang} -> {target_lang}, context: {context}")
            
            # バイトデータからワークブックを読み込み
            workbook = openpyxl.load_workbook(io.BytesIO(file_data))
            
            # 文脈に応じた前処理ルールを取得
            replacements = self.get_context_replacements(context)
            
            total_cells_translated = 0
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                logger.info(f"Processing sheet: {sheet_name}")
                
                # 翻訳対象のセルとテキストを収集
                cells_to_translate = []
                texts_to_translate = []
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.strip():
                            # 翻訳対象かどうかを判定
                            if self.should_translate_text(cell.value):
                                cells_to_translate.append(cell)
                                # 前処理を適用
                                processed_text = self.preprocess_text(cell.value, replacements)
                                texts_to_translate.append(processed_text)
                
                # 翻訳対象がある場合のみ翻訳実行
                if texts_to_translate:
                    logger.info(f"Translating {len(texts_to_translate)} cells in sheet {sheet_name}")
                    
                    # バッチサイズを制限して処理
                    batch_size = 50
                    for i in range(0, len(texts_to_translate), batch_size):
                        batch_texts = texts_to_translate[i:i + batch_size]
                        batch_cells = cells_to_translate[i:i + batch_size]
                        
                        # DeepL APIで翻訳
                        results = self.translator.translate_text(
                            batch_texts,
                            source_lang=source_lang,
                            target_lang=target_lang
                        )
                        
                        # 翻訳結果をセルに書き戻し
                        for cell, result in zip(batch_cells, results):
                            cell.value = result.text
                            total_cells_translated += 1
            
            logger.info(f"Translation completed: {total_cells_translated} cells translated")
            
            # 翻訳後のファイルをバイトデータとして返す
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.read()
            
        except deepl.exceptions.AuthorizationError:
            logger.error("DeepL API authorization error")
            raise Exception("DeepL APIキーが無効です。")
        except deepl.exceptions.QuotaExceededException:
            logger.error("DeepL API quota exceeded")
            raise Exception("DeepL APIの使用制限に達しました。")
        except Exception as e:
            logger.error(f"Translation error: {str(e)}")
            raise Exception(f"翻訳処理中にエラーが発生しました: {str(e)}")
    
    def _get_translation_context(self, context: str) -> str:
        """
        DeepL API用の翻訳コンテキストメッセージを生成
        
        Args:
            context: ユーザーが入力した文脈
            
        Returns:
            翻訳用コンテキストメッセージ
        """
        if not context:
            return "This is a general business document."
        
        context_lower = context.lower()
        
        if "日程" in context_lower or "itinerary" in context_lower:
            return "This is a travel itinerary for a tour group."
        elif "事業計画" in context_lower or "business plan" in context_lower:
            return "This is a business plan document with financial projections."
        elif "財務" in context_lower or "financial" in context_lower:
            return "This is a financial report or accounting document."
        elif "会議" in context_lower or "meeting" in context_lower:
            return "This is a meeting agenda or minutes document."
        else:
            return f"This is a {context} document."
    
    def validate_api_key(self) -> bool:
        """
        DeepL APIキーの有効性を検証
        
        Returns:
            APIキーが有効かどうか
        """
        try:
            # 簡単なテスト翻訳を実行
            result = self.translator.translate_text("テスト", source_lang="JA", target_lang="EN-US")
            logger.info(f"API key validation successful: {result.text}")
            return True
        except deepl.exceptions.AuthorizationError:
            logger.error("API key validation failed: Authorization error")
            return False
        except Exception as e:
            logger.error(f"API key validation failed: {str(e)}")
            return False